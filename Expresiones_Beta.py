# MIT License
# Copyright (c) 2024 Javier Chacón Milán.
# 
# Permission is hereby granted, free of charge, to any person obtaining a copy
# of this software and associated documentation files (the "Software"), to deal
# in the Software without restriction, including without limitation the rights
# to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
# copies of the Software, and to permit persons to whom the Software is
# furnished to do so, subject to the following conditions:
# 
# The above copyright notice and this permission notice shall be included in all
# copies or substantial portions of the Software.
# 
# THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
# IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
# FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
# AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
# LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
# OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
# SOFTWARE.


import cv2
import time
from fer import FER
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.drawing.image import Image  # Para insertar imágenes en Excel
import os

# Función para convertir segundos a formato HH:MM:SS
def convertir_a_hhmmss(segundos):
    horas = int(segundos // 3600)
    minutos = int((segundos % 3600) // 60)
    segundos = int(segundos % 60)
    return f"{horas:02d}:{minutos:02d}:{segundos:02d}"

# Inicializar el detector de emociones de la biblioteca fer
emotion_detector = FER()

# Inicializar la captura de video desde la webcam
cap = cv2.VideoCapture(0)

# Verificar si la cámara se abrió correctamente
if not cap.isOpened():
    print("No se pudo abrir la cámara")
    exit()

# Inicializar variables para almacenar los tiempos en listas
neutral_times = []
happy_times = []
disgust_times = []

# Inicializar variables para el cálculo de tiempo
emotion_times = {
    "neutral": 0,
    "happy": 0,
    "angry": 0,
    "disgust": 0,
}
start_time = time.time()  # Tiempo inicial
current_emotion = "neutral"
last_switch_time = start_time

while True:
    # Capturar frame por frame
    ret, frame = cap.read()

    if not ret:
        print("No se pudo recibir el frame (fin de transmisión)")
        break

    # Detectar emociones en la imagen
    emotions = emotion_detector.detect_emotions(frame)

    # Determinar la emoción predominante si se detecta algún rostro
    if emotions:
        # Obtener la emoción predominante
        dominant_emotion = max(emotions[0]["emotions"], key=emotions[0]["emotions"].get)
    else:
        dominant_emotion = "neutral"

    # Asignar color según la emoción predominante
    if dominant_emotion == "happy":
        color = (0, 255, 0)  # Verde (sonrisa)
        emotion_label = "happy"
    elif dominant_emotion == "angry":
        color = (0, 0, 255)  # Rojo (ira o ceño fruncido)
        emotion_label = "angry"
    elif dominant_emotion == "disgust":
        color = (0, 165, 255)  # Naranja (desagrado)
        emotion_label = "disgust"
    else:
        color = (255, 0, 0)  # Azul (neutral)
        emotion_label = "neutral"

    # Cambiar de estado si la emoción cambia
    if emotion_label != current_emotion:
        current_time = time.time()
        elapsed_time = current_time - last_switch_time  # Tiempo transcurrido en la emoción anterior
        emotion_times[current_emotion] += elapsed_time  # Acumular el tiempo de la emoción previa
        # Momento desde el inicio del programa en segundos
        moment_since_start = current_time - start_time
        
        # Guardar el tiempo en el vector correspondiente según la emoción detectada
        if current_emotion == "neutral":
            neutral_times.append(convertir_a_hhmmss(moment_since_start))
        elif current_emotion == "happy":
            happy_times.append(convertir_a_hhmmss(moment_since_start))
        elif current_emotion == "angry" or current_emotion == "disgust":
            disgust_times.append(convertir_a_hhmmss(moment_since_start))
        
        # Actualizar el tiempo del último cambio
        last_switch_time = current_time
        # Actualizar la emoción actual
        current_emotion = emotion_label

    # Colocar etiqueta en español
    if emotion_label == "happy":
        etiqueta = "sonrisa"
    elif emotion_label == "angry" or emotion_label == "disgust":
        etiqueta = "rechazo"
    else:
        etiqueta = "normal"

    # Dibujar un rectángulo en la imagen según la emoción detectada
    for emotion in emotions:
        (x, y, w, h) = emotion["box"]
        cv2.rectangle(frame, (x, y), (x + w, y + h), color, 2)
        cv2.putText(frame, etiqueta, (x, y - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.9, color, 2)

    # Mostrar el frame con las detecciones
    cv2.imshow('Detector de emociones', frame)

    # Salir del bucle al presionar la tecla 'q'
    if cv2.waitKey(1) == ord('q'):
        # Registrar el tiempo final antes de salir
        current_time = time.time()
        emotion_times[current_emotion] += current_time - last_switch_time  # Registrar el tiempo de la última emoción
        break

# Liberar el objeto de captura y cerrar todas las ventanas
cap.release()
cv2.destroyAllWindows()

# Imprimir los resultados
total_time = time.time() - start_time

print(f"Tiempo total con sonrisa: {convertir_a_hhmmss(emotion_times['happy'])}")
print(f"Tiempo total neutral: {convertir_a_hhmmss(emotion_times['neutral'])}")
print(f"Tiempo total con desagrado e ira: {convertir_a_hhmmss(emotion_times['angry'] + emotion_times['disgust'])}")
print(f"Tiempo total de la sesión: {convertir_a_hhmmss(total_time)}")

# Imprimir tiempos individuales almacenados en los vectores con formato HH:MM:SS
print("\n\nTiempos de cambios a estado 'neutral':", neutral_times)
print("Tiempos de cambios a estado 'sonrisa':", happy_times)
print("Tiempos de cambios a estado 'rechazo (angry/disgust)':", disgust_times, "\n")

# Preparar datos para el gráfico circular
labels = []
times = []
colors = []

# Comprobamos las emociones que se han detectado en la ejecución para el gráfico circular
if emotion_times['neutral'] > 0:
    labels.append('Neutral')
    times.append(emotion_times['neutral'])
    colors.append('blue')

if emotion_times['happy'] > 0:
    labels.append('Sonrisa')
    times.append(emotion_times['happy'])
    colors.append('green')

if emotion_times['angry'] > 0 or emotion_times['disgust'] > 0:
    labels.append('Rechazo')
    times.append(emotion_times['angry'] + emotion_times['disgust'])
    colors.append('orange')

# Crear gráfico circular y guardarlo como imagen
plt.figure(figsize=(4, 4))
plt.pie(times, labels=labels, autopct='%1.1f%%', colors=colors, startangle=90)
plt.title('Distribución de tiempo por emoción')
plt.axis('equal')  # Para hacer el gráfico circular

# Guardar la imagen del gráfico
if not os.path.exists("./documentos"):
    os.makedirs("./documentos")
plt.savefig("./documentos/emotion_pie_chart.png")
plt.show()

# Exportar los datos a un archivo Excel
wb = Workbook()
ws = wb.active


# Estilo de colores para las celdas de cabecera
header_fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")  # Rosa
neutral_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")  # Azul
happy_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")    # Verde
disgust_fill = PatternFill(start_color="FFA07A", end_color="FFA07A", fill_type="solid")  # Naranja

# Estilo de colores para las celdas de contenido
content_neutral_fill= PatternFill(start_color="CCECFF", end_color="CCECFF", fill_type="solid")  # Azul Claro
content_happy_fill= PatternFill(start_color="D1FFD1", end_color="D1FFD1", fill_type="solid")  # Verde Claro
content_disgust_fill= PatternFill(start_color="FFCC99", end_color="FFCC99", fill_type="solid")  # Naranja Claro

# Títulos para las columnas
ws['A1'] = 'Tiempo Sesion'
ws['B1'] = 'Total Neutral'
ws['C1'] = 'Total Sonrisa'
ws['D1'] = 'Total Desagrado/Ira'

# Colocar el tiempo total de la sesión y los totales de cada emoción
ws['A2'] = convertir_a_hhmmss(total_time)
ws['B2'] = convertir_a_hhmmss(emotion_times['neutral'])
ws['C2'] = convertir_a_hhmmss(emotion_times['happy'])
ws['D2'] = convertir_a_hhmmss(emotion_times['angry'] + emotion_times['disgust'])


# Colocar colores de los tiempos de cada emoción
ws['B2'].fill = content_neutral_fill
ws['C2'].fill = content_happy_fill
ws['D2'].fill = content_disgust_fill

# Etiquetas para los tiempos individuales
ws['A4'] = 'Tiempos Neutral'
ws['B4'] = 'Tiempos Sonrisa'
ws['C4'] = 'Tiempos Desagrado'

# Colores para los tiempos individuales
ws['A4'].fill = neutral_fill
ws['B4'].fill = happy_fill
ws['C4'].fill = disgust_fill


# Rellenar los tiempos individuales en las columnas correspondientes
for i, time_value in enumerate(neutral_times, start=5):
    ws[f'A{i}'] = time_value


for i, time_value in enumerate(happy_times, start=5):
    ws[f'B{i}'] = time_value


for i, time_value in enumerate(disgust_times, start=5):
    ws[f'C{i}'] = time_value



# Aplicar colores
ws['A1'].fill = header_fill
ws['B1'].fill = neutral_fill
ws['C1'].fill = happy_fill
ws['D1'].fill = disgust_fill

# Insertar la imagen del gráfico en el archivo Excel
img = Image("./documentos/emotion_pie_chart.png")
ws.add_image(img, 'E5')  # Insertar la imagen a partir de la celda E5

# Guardar el archivo en una carpeta 'documentos' en el directorio actual
wb.save("./documentos/emotion_times.xlsx")

print("Datos y gráfico exportados exitosamente a emotion_times.xlsx")